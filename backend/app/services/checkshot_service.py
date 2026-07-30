"""
services/checkshot_service.py
--------------------------------
Parses and stores a real checkshot / time-depth-survey workbook (one sheet
per well, columns including TWT(ms) and Depth(m)) -- e.g. Zamzama-TDS.xlsx
-- so well_seismic_tie.checkshot_anchor_shift / direct_tie_service can
anchor the well-to-seismic tie to real checkshot stations instead of
falling back to the full +/-100ms statistical search everywhere.

Column layout (verified against the real uploaded workbook, not just
assumed from the reference pipeline's comment): each sheet's data rows are
(index, label, TWT(ms), <blank>, Depth(m)) starting at row 2 (row 1 is
either a header or empty depending on the sheet) -- column C (index 2) is
TWT(ms), column E (index 4) is Depth(m). Sheets are stored keyed by their
own name (usually the field's base well name, e.g. 'Z-02'); lookups
tolerate this app's well_id having an extra suffix the sheet name doesn't
(e.g. well_id 'Z-02_RAW' from the LAS filename vs. sheet 'Z-02') -- see
_find_checkshot_record.
"""

from __future__ import annotations

import io

from app.checkshot_repository import CheckshotRecord, get_checkshot_repository

TWT_COLUMN_INDEX = 2
DEPTH_COLUMN_INDEX = 4


class CheckshotValidationError(Exception):
    """Raised when an uploaded checkshot workbook can't be parsed at all."""


def parse_checkshot_workbook(file_bytes: bytes) -> dict[str, list[tuple[float, float]]]:
    """Returns {well_id: [(depth_m, twt_ms), ...]} sorted ascending by
    depth, one entry per sheet with at least one valid (TWT, Depth) row.
    Sheets/rows with unparseable or missing values are skipped, not fatal
    -- a checkshot survey commonly has sparse/partial coverage per well."""
    try:
        import openpyxl
    except ImportError as exc:
        raise CheckshotValidationError(
            "openpyxl is required to read checkshot workbooks (pip install openpyxl)."
        ) from exc

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001 -- any corrupt/non-xlsx file lands here
        raise CheckshotValidationError(f"Could not read workbook: {exc}") from exc

    checkshots: dict[str, list[tuple[float, float]]] = {}
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        points: list[tuple[float, float]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= max(TWT_COLUMN_INDEX, DEPTH_COLUMN_INDEX):
                continue
            twt, depth = row[TWT_COLUMN_INDEX], row[DEPTH_COLUMN_INDEX]
            if twt is None or depth is None:
                continue
            try:
                points.append((float(depth), float(twt)))
            except (TypeError, ValueError):
                continue
        if points:
            checkshots[sheet_name] = sorted(points, key=lambda p: p[0])

    if not checkshots:
        raise CheckshotValidationError(
            "No valid (Depth, TWT) checkshot points found in any sheet -- expected one sheet per "
            "well with TWT(ms) in column C and Depth(m) in column E."
        )
    return checkshots


def store_checkshot_workbook(file_bytes: bytes) -> dict[str, int]:
    """Parses and persists the workbook, one CheckshotRecord per well_id
    (sheet name). Returns {well_id: n_points} for the caller to report
    back. Overwrites any existing checkshot record for a well_id that
    reappears in this upload."""
    parsed = parse_checkshot_workbook(file_bytes)
    repo = get_checkshot_repository()
    counts: dict[str, int] = {}
    for well_id, points in parsed.items():
        repo.save(CheckshotRecord(well_id=well_id, points=[[d, t] for d, t in points]))
        counts[well_id] = len(points)
    return counts


def _find_checkshot_record(well_id: str):
    """Exact match first; then a prefix match against every stored
    record, tolerant of a suffix the app's well_id has that the
    checkshot sheet name doesn't (e.g. this app derives well_id from the
    LAS FILENAME -- las_loader._well_id_from_filename -- which can be
    'Z-02_RAW', while a checkshot workbook typically names its sheets
    after the field's base well name, 'Z-02', matching the LAS header's
    WELL mnemonic instead). A record's well_id only matches as a prefix
    if it's followed by a non-alphanumeric boundary (e.g. 'Z-02_RAW'
    matches stored 'Z-02', but stored 'Z-0' does NOT match 'Z-02_RAW' --
    the next character '2' is alphanumeric, not a real boundary), so
    'Z-02' and 'Z-020' can't cross-match each other. Case-insensitive,
    since las_loader uppercases well_id but an uploaded workbook's sheet
    names are used as-is."""
    repo = get_checkshot_repository()
    well_id_upper = well_id.upper()

    exact = repo.get(well_id)
    if exact is not None:
        return exact

    best = None
    for record in repo.list_all():
        candidate = record.well_id.upper()
        if candidate == well_id_upper:
            return record
        if well_id_upper.startswith(candidate) and not well_id_upper[len(candidate)].isalnum():
            if best is None or len(candidate) > len(best.well_id):
                best = record
    return best


def get_checkshot_points(well_id: str) -> list[tuple[float, float]]:
    """[(depth_m, twt_ms), ...] sorted ascending by depth, or [] if no
    checkshot data has been uploaded for this well -- NOT an error, since
    most callers should just fall back to the statistical tie search."""
    record = _find_checkshot_record(well_id)
    if record is None:
        return []
    return [(float(d), float(t)) for d, t in record.points]


def get_checkshot_status() -> dict[str, int]:
    """{well_id: n_points} for every well with uploaded checkshot data --
    backs a status endpoint so the frontend can show which wells have real
    checkshot coverage vs. which will use the statistical fallback."""
    return {record.well_id: len(record.points) for record in get_checkshot_repository().list_all()}
